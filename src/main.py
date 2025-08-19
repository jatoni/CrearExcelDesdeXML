import pandas as pd
import os
import xmltodict
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import pprint

registro_alumnos = {}

def readXMLAndBuildData(archivoXML, isPaginate):
    dataBuilded = {}
    base_path = archivoXML
    carpetas = os.listdir(base_path)

    max_alumnos_por_hoja = 26

    for carpeta in carpetas:
        ruta_carpeta = os.path.join(base_path, carpeta)
        if os.path.isdir(ruta_carpeta):
            archivos = sorted(os.listdir(ruta_carpeta)) 
            name_file_base = os.path.basename(carpeta)
            if isPaginate:
                dataBuilded = buildDataPaginate(archivos, name_file_base, max_alumnos_por_hoja, dataBuilded, ruta_carpeta)
            else:
                dataBuilded = buildDataWithoutPaginate(archivos, name_file_base, dataBuilded, ruta_carpeta)
    return dataBuilded

def buildDataWithoutPaginate(archivos, name_file_base, dataBuilded, ruta_carpeta):
    for archivo in archivos:
        if not archivo.endswith('.xml'):
            continue
        name_file = f"{name_file_base}"
        ruta_archivo = os.path.join(ruta_carpeta, archivo)
        folio_Control = os.path.splitext(os.path.basename(archivo))[0]

        with open(ruta_archivo, 'r', encoding='utf-8') as f:
            contenido_xml = f.read()

        diccionarioData = xmltodict.parse(contenido_xml)["Dec"]
        #pprint.pprint(diccionarioData)
        alumno = diccionarioData["Alumno"]

        partes = [
            alumno.get("@nombre", ""),
            alumno.get("@primerApellido", ""),
            alumno.get("@segundoApellido", "")
        ]   
        nombreAlumno = " ".join(p for p in partes if p)
        
        curp = diccionarioData["Alumno"]["@curp"]
        carrera = diccionarioData["Carrera"]["@nombreCarrera"]
        name_clave = diccionarioData["Carrera"]["@claveCarrera"]

        registro = {
            "ALUMNO": nombreAlumno,
            "CURP": curp,
            "Carrera": carrera,
            "CLAVE_DE_CARRERA": name_clave,
            "FOLIO_DE_CONTROL": folio_Control
        }
        dataBuilded.setdefault(name_file, []).append(registro)
    return dataBuilded

def buildDataPaginate(archivos, name_file_base, max_alumnos_por_hoja, dataBuilded, ruta_carpeta):
    hoja_num = 1
    for archivo in archivos:
        if not archivo.endswith('.xml'):
            continue
        name_file = f"{name_file_base} Hoja {hoja_num}"
        if len(dataBuilded.get(name_file, [])) >= max_alumnos_por_hoja:
            hoja_num += 1
            name_file = f"{name_file_base} Hoja {hoja_num}"

        ruta_archivo = os.path.join(ruta_carpeta, archivo)
        folio_Control = os.path.splitext(os.path.basename(archivo))[0]

        with open(ruta_archivo, 'r', encoding='utf-8') as f:
            contenido_xml = f.read()

        diccionarioData = xmltodict.parse(contenido_xml)["TituloElectronico"]

        nombreAlumno = (
            diccionarioData["Profesionista"]["@nombre"] + " " +
            diccionarioData["Profesionista"]["@primerApellido"] + " " +
            diccionarioData["Profesionista"]["@segundoApellido"]
        )
        curp = diccionarioData["Profesionista"]["@curp"]
        programa = diccionarioData["Carrera"]["@nombreCarrera"]
        lugar_expedicion = diccionarioData["Expedicion"]["@entidadFederativa"]
        fecha_expedicion = diccionarioData["Expedicion"]["@fechaExpedicion"]
        rvoe = diccionarioData["Carrera"]["@numeroRvoe"]
        clave_insitucion = diccionarioData["Institucion"]["@cveInstitucion"]
        name_clave = diccionarioData["Carrera"]["@cveCarrera"]

        registro = {
            "NUM_PROG": len(dataBuilded.get(name_file, [])) + 1,
            "ALUMNO": nombreAlumno,
            "CURP": curp,
            "PROGRAMA": programa,
            "CLAVE_DE_CARRERA": name_clave,
            "FOLIO_DE_CONTROL": folio_Control,
            "LUGAR_DE_EXPEDICION": lugar_expedicion,
            "FECHA_DE_EXPEDICION": fecha_expedicion,
            "RVOE": rvoe,
            "CLAVE_DE_INSTITUCION": clave_insitucion,
            "FOLIO_DIGITAL": diccionarioData["Autenticacion"]["@folioDigital"],
        }
    return dataBuilded.setdefault(name_file, []).append(registro)

def agregar_hoja_nueva_excel(ruta_excel, dic):
    coutn = 0
    libro = load_workbook(ruta_excel)
    hoja_base = libro[libro.sheetnames[0]]  # Primera hoja como plantilla

    for name_file, info in dic.items():
        if name_file in libro.sheetnames:
            print(f"La hoja '{name_file}' ya existe. No se agregará hoja nueva.")
            continue

        # Copiar hoja base
        nueva_hoja = libro.copy_worksheet(hoja_base)
        nueva_hoja.title = name_file

        df_info = pd.DataFrame(info)

        # ---------- TABLA DE DATOS ----------
        start_row = 15  # fila donde empieza tu tabla en la plantilla
        for r_idx, row in enumerate(dataframe_to_rows(df_info, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                nueva_hoja.cell(row=r_idx, column=c_idx, value=value)

        # ---------- FECHA ----------
        fecha_valor = ""
        if "FECHA_DE_EXPEDICION" in df_info.columns and not df_info["FECHA_DE_EXPEDICION"].empty:
            fecha_valor = str(df_info["FECHA_DE_EXPEDICION"].iloc[0])

        # ---------- HOJA ----------
        numero_hoja = int(re.search(r'\d+$', name_file).group())
        prefijo_paq = re.match(r'PAQ\. T-\d+', name_file).group()
        total_hojas = sum(1 for nombre in dic.keys() if nombre.startswith(prefijo_paq))

        # Colocar fecha y hoja en celdas fijas
        nueva_hoja["J8"] = fecha_valor
        nueva_hoja["J9"] = f"{numero_hoja}/{total_hojas}"

        # ---------- FIRMAS ----------
        responsable = " "
        if "CLAVE_DE_INSTITUCION" in df_info.columns:
            if (df_info["CLAVE_DE_INSTITUCION"].astype(str) == "150901").any():
                responsable = "Mtro. Luis Ernesto Gutiérrez Martínez"
            else:
                responsable = "Mtra. Lilibeth Hernandez Alva"
        
        
        nueva_hoja.merge_cells("C51:D51")
        
        nueva_hoja["C51"] = responsable

        
        nueva_hoja["I51"] = "Mtra. Dely Karolina Urbano Sanchez"
        nueva_hoja["I52"] = "RECTOR"

        print(f"Hoja '{name_file}' creada con fecha, hoja y firmas en posiciones exactas.")
        coutn += 1
    print(f"Total de hojas agregadas: {coutn}")
    libro.save(ruta_excel)
    
def obtenerDatosAlumnos(data):
    print("Obteniendo Datos de Alumnos...")
    ruta = "./DocuemntosTitulos/alumnos.xlsx"
    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        for nombre_hoja, lista_dicts in data.items():
            df = pd.DataFrame(lista_dicts)
            df.to_excel(writer, sheet_name=nombre_hoja, index=False)

    

opc = 0
while opc != 3:
    opc = int(input("Menú \n1.- Generar Folio de titulos\n2.-Obtener datos de alumnos y enlistarlos\n3.-Salir\nChoose one: "))
    if opc == 1:
        ruta = "/Users/juanantoniotorres/Documents/ProyectoSalvandoVidaJacqueline/DocuemntosTitulos/Libro de Control de Folios de Titulos y Grados Electrónicos 2025.xlsx"
        registro_alumnos = readXMLAndBuildData("./ArchivosXML", True)
        dic = registro_alumnos
        agregar_hoja_nueva_excel(ruta, dic)
    elif opc == 2:
        registro_alumnos = readXMLAndBuildData("./ArchivosXMLObtenerDatos", False)
        obtenerDatosAlumnos(registro_alumnos)